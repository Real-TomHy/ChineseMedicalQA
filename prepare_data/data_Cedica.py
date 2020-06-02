import requests
from lxml import html
import pandas
from openpyxl import Workbook
import re

class zhongyao():
    def __init__(self):
        self.text_all = dict()
        self.url="http://www.a-hospital.com/w/%E6%9C%AC%E8%8D%89%E7%BA%B2%E7%9B%AE"
        self.headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36"}
    def get_parse_html(self,url): #将网页源代码转换成xpath对象的函数
        res=requests.get(url,headers=self.headers)
        res.encoding = "utf-8"
        html_text = res.text
        parse_html = html.etree.HTML(html_text)#将网页源代码转换成xpath对象
        return parse_html

    def get_url(self):
        res=requests.get(self.url,headers=self.headers)
        res.encoding="utf-8"
        en = html.etree.HTML(res.text)
        text_title=en.xpath('//h3/span[@class="mw-headline"]/text()')
        text_url1 = en.xpath('//*[@id="bodyContent"]/p[5]/a/@href')
        self.get_nate(text_url1)
        text_url2 = en.xpath('//*[@id="bodyContent"]/p[6]/a/@href')
        text_url3 = en.xpath('//*[@id="bodyContent"]/p[7]/a/@href')
        # self.get_nate(text_url3)
        text_url4 = en.xpath('//*[@id="bodyContent"]/p[8]/a/@href')
        #self.get_nate(text_url4)
        text_url5 = en.xpath('//*[@id="bodyContent"]/p[9]/a/@href')
        #self.get_nate(text_url5)
        text_url6 = en.xpath('//*[@id="bodyContent"]/p[10]/a/@href')
        # self.get_nate(text_url6)
        text_url7 = en.xpath('//*[@id="bodyContent"]/p[11]/a/@href')
        # self.get_nate(text_url7)
        text_url8 = en.xpath('//*[@id="bodyContent"]/p[12]/a/@href')
        # self.get_nate(text_url8)
        text_url9 = en.xpath('//*[@id="bodyContent"]/p[13]/a/@href')
        # self.get_nate(text_url9)
        text_url10 = en.xpath('//*[@id="bodyContent"]/p[14]/a/@href')
        #self.get_nate(text_url10)
        text_url11 = en.xpath('//*[@id="bodyContent"]/p[15]/a/@href')
        # self.get_nate(text_url11)
        text_url12 = en.xpath('//*[@id="bodyContent"]/p[16]/a/@href')
        # self.get_nate(text_url12)
        text_url13 = en.xpath('//*[@id="bodyContent"]/p[17]/a/@href')
        # self.get_nate(text_url13)
        text_url14 = en.xpath('//*[@id="bodyContent"]/p[18]/a/@href')
        # self.get_nate(text_url14)
        text_url15 = en.xpath('//*[@id="bodyContent"]/p[19]/a/@href')
        # self.get_nate(text_url15)
    def get_text_alias(self,parse_html_text):
        try:
            alias_data=""
            text_alias = parse_html_text.xpath('//*[@id="bodyContent"]/p[1]/text()')
            text_alias = ''.join(text_alias)
            alias_data = text_alias.split('」')[1]
        except:
            return ""
        return alias_data

    def get_text_smell(self, parse_html_text):
        try:
            smell_data = ""
            text_smell = parse_html_text.xpath('//*[@id="bodyContent"]/p[2]/text()')
            text_smell = ''.join(text_smell)
            smell_data = text_smell.split('」')[1]
        except:
            return ""
        return smell_data
    def get_text_cure(self, parse_html_text):
        try:
            new_cure = ""
            text_cure1 = parse_html_text.xpath('//*[@id="bodyContent"]/p[3]/text()')
            text_cure2 = parse_html_text.xpath('string(//*[@id="bodyContent"]/p[4])')
            text_cure3 = parse_html_text.xpath('string(//*[@id="bodyContent"]/p[5])')
            text_cure4 = parse_html_text.xpath('string(//*[@id="bodyContent"]/p[6])')
            text_cure1 = ''.join(text_cure1)
            text_cure2 = ''.join(text_cure2)
            text_cure3 = ''.join(text_cure3)
            text_cure4 = ''.join(text_cure4)
            new_cure = text_cure1 + text_cure2 + text_cure3 + text_cure4
            new_cure = new_cure.split('」')[1]
        except:
            return ""
        return new_cure
    # def save(self,row):
    #     for i in row:
    #         with open('土部.xlsx', "") as f:
    #             f.write(i)
    def get_nate(self,text_url):
        count=0
        rows=[]
        for link in text_url:#对所有帖子的站内链接进行遍历 拼接完整的帖子链接
            t_url="http://www.a-hospital.com"+link#拼接得到帖子的url
            parse_html_text=self.get_parse_html(t_url)
            text_name = parse_html_text.xpath('//*[@id="firstHeading"]/text()')
            text_name = ''.join(text_name)
            text_name = text_name.split('/')[1]  # 取‘/’右边的
            text_alias=self.get_text_alias(parse_html_text)
            text_smell=self.get_text_smell(parse_html_text)
            text_cure=self.get_text_cure(parse_html_text)
            # try:
            #     text_alias = parse_html_text.xpath('//*[@id="bodyContent"]/p[1]/text()')
            #     text_alias = ''.join(text_alias)
            #     text_alias = text_alias.split('」')[1]
            # except:
            #     return text_alias
            # try:
            #     text_smell = parse_html_text.xpath('//*[@id="bodyContent"]/p[2]/text()')
            #     text_smell = ''.join(text_smell)
            #     text_smell = text_smell.split('」')[1]
            # except:
            #     text_smell=""
            #
            #
            # try:
            #     text_cure1 = parse_html_text.xpath('//*[@id="bodyContent"]/p[3]/text()')
            #     text_cure2 = parse_html_text.xpath('string(//*[@id="bodyContent"]/p[4])')
            #     text_cure3 = parse_html_text.xpath('string(//*[@id="bodyContent"]/p[5])')
            #     text_cure4 = parse_html_text.xpath('string(//*[@id="bodyContent"]/p[6])')
            #     text_cure1=''.join(text_cure1)
            #     text_cure2=''.join(text_cure2)
            #     text_cure3=''.join(text_cure3)
            #     text_cure4=''.join(text_cure4)
            #     new_cure = text_cure1 + text_cure2 + text_cure3 + text_cure4
            #     new_cure = new_cure.split('」')[1]
            # except Exception:
            #     return new_cure #对于某些没有主治的异常处理

            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'name'
            ws['B1'] = 'alias'
            ws['C1'] = 'smell'
            ws['D1'] = 'cure'
            row=[text_name,text_alias,text_smell,text_cure]

            rows.append(row)
            #self.save(row)
            # print(row)
            # ws.append(row)
            for new_row in rows:
                ws.append(new_row)
            count+=1
            print(count)
            wb.save('草部.xlsx')
            # frame = pandas.DataFrame(columns=['name','alias', 'smell','cure'])
            # frame['name'] = text_name
            # frame['alias'] = text_alias
            # frame['smell'] = text_smell
            # frame['cure'] = new_cure
            # frame.to_excel('./data/草部.xlsx')
            # text_cure3= parse_html_text.xpath('string(//*[@id="bodyContent"])')
            # a="""「主治」"""
            # b="""参考"""
            # new_cure = re.search('^1.*?b$',text_cure3,re.S)
            #print(row)

zhongyao = zhongyao()
zhongyao.get_url()